from fastapi import FastAPI, Request, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, StreamingResponse
import json, os, io
from datetime import datetime
from models.schemas import RiskAssessmentResponse, RuleDefinition
from models.scorer import load_config, save_config, calculate_risk
from models import database as db

app = FastAPI(title="宝安区欠薪预警API", description="穿透式前瞻性欠薪预警引擎", version="2.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

CONFIG_PATH = os.path.join(os.path.dirname(__file__), 'rules_config.json')
STATIC_PATH = os.path.join(os.path.dirname(__file__), 'static')
os.makedirs(STATIC_PATH, exist_ok=True)
app.mount("/static", StaticFiles(directory=STATIC_PATH), name="static")

# === 前端 ===
@app.get("/", response_class=HTMLResponse, tags=["前端应用"])
async def serve_frontend():
    p = os.path.join(STATIC_PATH, 'index.html')
    if not os.path.exists(p): return HTMLResponse("<h1>前端页面尚未部署</h1>", status_code=404)
    with open(p, 'r', encoding='utf-8') as f: return f.read()

# === 认证 ===
@app.post("/api/v1/auth/login", tags=["认证"])
async def login(request: Request):
    data = await request.json()
    config = load_config()
    auth = config.get("system", {}).get("auth", {})
    if data.get("username") == auth.get("username", "admin") and data.get("password") == auth.get("password", "admin123"):
        return {"status": "success", "message": "登录成功", "username": data["username"]}
    raise HTTPException(401, "用户名或密码错误")

# === 系统配置 ===
@app.get("/api/v1/config/system", tags=["系统配置"])
async def get_system_config():
    cfg = load_config().get("system", {})
    # flatten risk_thresholds for frontend compatibility
    if "risk_thresholds" in cfg:
        t = cfg.pop("risk_thresholds", {})
        cfg["risk_red_threshold"] = t.get("red", 20)
        cfg["risk_yellow_threshold"] = t.get("yellow", 16)
        cfg["risk_blue_threshold"] = t.get("blue", 11)
    # remove auth config from response
    cfg.pop("auth", None)
    return cfg

@app.put("/api/v1/config/system", tags=["系统配置"])
async def update_system_config(request: Request):
    data = await request.json(); c = load_config()
    # unflatten risk_thresholds
    if "risk_red_threshold" in data:
        data["risk_thresholds"] = {
            "red": data.pop("risk_red_threshold", 20),
            "yellow": data.pop("risk_yellow_threshold", 16),
            "blue": data.pop("risk_blue_threshold", 11)
        }
    # preserve auth config
    data["auth"] = c.get("system", {}).get("auth", {})
    c["system"] = data; save_config(c)
    return {"message": "系统配置更新成功", "status": "success"}

# === 规则 CRUD ===
@app.get("/api/v1/config/{domain}", tags=["规则管理"])
async def get_domain_rules(domain: str):
    c = load_config().get(domain)
    if c is None: raise HTTPException(404, f"领域 '{domain}' 不存在")
    return c

@app.get("/api/v1/config/{domain}/{key}", tags=["规则管理"])
async def get_indicator(domain: str, key: str):
    c = load_config().get(domain)
    if c is None: raise HTTPException(404)
    for r in c.get("rules", []):
        if r.get("key") == key: return r
    raise HTTPException(404, f"指标 '{key}' 不存在")

@app.post("/api/v1/config/{domain}", tags=["规则管理"])
async def add_indicator(domain: str, rule: RuleDefinition):
    c = load_config()
    if domain not in c: raise HTTPException(404)
    rules = c[domain].get("rules", [])
    for e in rules:
        if e.get("key") == rule.key: raise HTTPException(409, f"指标标识 '{rule.key}' 已存在")
    rules.append(rule.model_dump()); c[domain]["rules"] = rules; save_config(c)
    return {"message": f"指标 '{rule.name}' 新增成功", "status": "success"}

@app.put("/api/v1/config/{domain}/{key}", tags=["规则管理"])
async def update_indicator(domain: str, key: str, rule: RuleDefinition):
    c = load_config()
    if domain not in c: raise HTTPException(404)
    rules = c[domain].get("rules", [])
    for i, e in enumerate(rules):
        if e.get("key") == key:
            rules[i] = rule.model_dump(); c[domain]["rules"] = rules; save_config(c)
            return {"message": f"指标 '{rule.name}' 更新成功", "status": "success"}
    raise HTTPException(404)

@app.delete("/api/v1/config/{domain}/{key}", tags=["规则管理"])
async def delete_indicator(domain: str, key: str):
    c = load_config()
    if domain not in c: raise HTTPException(404)
    rules = c[domain].get("rules", [])
    new_rules = [r for r in rules if r.get("key") != key]
    if len(new_rules) == len(rules): raise HTTPException(404)
    c[domain]["rules"] = new_rules; save_config(c)
    return {"message": f"指标 '{key}' 已删除", "status": "success"}

# === 动态预测 ===
@app.post("/api/v1/predict/{domain}", response_model=RiskAssessmentResponse, tags=["预警模型分析"])
async def predict_risk(domain: str, request: Request):
    if domain not in load_config(): raise HTTPException(404)
    data = await request.json()
    return calculate_risk(domain, data)

# === 企业管理 ===
@app.get("/api/v1/enterprises", tags=["企业管理"])
async def list_enterprises(risk_level: str = None, enterprise_type: str = None, search: str = None, district: str = None, page: int = 1, page_size: int = 50):
    return db.list_enterprises(risk_level, enterprise_type, search, district, page, page_size)

@app.post("/api/v1/enterprises", tags=["企业管理"])
async def create_enterprise(request: Request):
    data = await request.json()
    if not data.get('name') or not data.get('credit_code'):
        raise HTTPException(400, "企业名称和信用代码为必填项")
    try:
        eid = db.create_enterprise(data)
        return {"message": "企业创建成功", "id": eid, "status": "success"}
    except Exception as e:
        raise HTTPException(409, f"创建失败: {str(e)}")

# === 数据导出（必须在 {eid} 路由之前注册以避免路径冲突） ===
@app.get("/api/v1/enterprises/export", tags=["数据导出"])
async def export_enterprises(risk_level: str = None):
    """
    按风险等级导出企业数据（CSV格式）。
    risk_level: 红色预警/黄色预警/蓝色预警/绿色预警/all（默认全部）
    """
    data = db.export_enterprises_by_level(risk_level)

    # 生成XLSX
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "企业风险数据"
    
    headers = ['企业名称', '统一社会信用代码', '企业类型', '所在街道', '所属行业',
               '员工数', '风险分数', '风险等级', '是否红线', '最后评估时间', '建议处置']
    ws.append(headers)

    for row in data:
        ws.append([
            row.get('name', ''),
            row.get('credit_code', ''),
            row.get('enterprise_type', ''),
            row.get('district', ''),
            row.get('industry', ''),
            row.get('employee_count', ''),
            row.get('risk_score', ''),
            row.get('risk_level', ''),
            '是' if row.get('is_red_line') else '否',
            row.get('last_assessed_at', ''),
            row.get('recommended_actions', '')
        ])

    safe_level = (risk_level or 'all')
    if safe_level == '红色预警': safe_level = 'red'
    elif safe_level == '黄色预警': safe_level = 'yellow'
    elif safe_level == '蓝色预警': safe_level = 'blue'
    elif safe_level == '绿色预警': safe_level = 'green'
    filename = f"wage_arrears_risk_{safe_level}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/api/v1/enterprises/{eid}", tags=["企业管理"])
async def get_enterprise(eid: int):
    ent = db.get_enterprise(eid)
    if not ent: raise HTTPException(404, "企业不存在")
    return ent

@app.put("/api/v1/enterprises/{eid}", tags=["企业管理"])
async def update_enterprise(eid: int, request: Request):
    data = await request.json()
    db.update_enterprise(eid, data)
    return {"message": "企业信息更新成功", "status": "success"}

@app.delete("/api/v1/enterprises/{eid}", tags=["企业管理"])
async def delete_enterprise(eid: int):
    db.delete_enterprise(eid)
    return {"message": "企业已删除", "status": "success"}

@app.post("/api/v1/enterprises/{eid}/assess", tags=["企业管理"])
async def assess_enterprise(eid: int, request: Request):
    ent = db.get_enterprise(eid)
    if not ent: raise HTTPException(404, "企业不存在")
    data = await request.json()
    domain = ent['enterprise_type']
    result = calculate_risk(domain, data)
    details = [d.model_dump() for d in result.details]
    db.save_assessment(eid, result.total_score, result.risk_level, result.is_red_line_triggered, details, result.recommended_actions, data)
    return {"message": "评估完成", "status": "success", "result": result}

@app.post("/api/v1/enterprises/push-indicators", tags=["企业管理"], summary="按统一信用代码推送指标并自动评估")
async def push_indicators(request: Request):
    """
    核心数据对接接口：按统一社会信用代码推送指标数据，系统自动：
    1. 将指标数据写入 indicator_snapshots 快照表
    2. 按 credit_code 关联到对应企业
    3. 立即触发风险评估，将结果写回企业档案及 assessment_history 历史表

    请求体格式：
    {
      "credit_code": "91440306XXXXXXXX",
      "data_source": "社保局",
      "indicators": { ...指标字段... }
    }
    """
    body = await request.json()
    credit_code = body.get('credit_code', '').strip()
    indicator_data = body.get('indicators', {})
    data_source = body.get('data_source', '外部系统推送')

    if not credit_code:
        raise HTTPException(400, "credit_code 为必填项")
    if not indicator_data:
        raise HTTPException(400, "indicators 指标数据不能为空")

    # 查找关联企业
    ent = db.get_enterprise_by_credit_code(credit_code)
    if not ent:
        raise HTTPException(404, f"未找到统一信用代码为 '{credit_code}' 的企业，请先在企业库中录入")

    # 写入指标快照
    db.push_indicator_snapshot(credit_code, ent['enterprise_type'], indicator_data, data_source)

    # 触发风险评估
    result = calculate_risk(ent['enterprise_type'], indicator_data)
    details = [d.model_dump() for d in result.details]
    db.save_assessment(ent['id'], result.total_score, result.risk_level, result.is_red_line_triggered, details, result.recommended_actions, indicator_data)

    return {
        "message": f"企业 '{ent['name']}' 指标推送与评估完成",
        "status": "success",
        "credit_code": credit_code,
        "enterprise_name": ent['name'],
        "risk_score": result.total_score,
        "risk_level": result.risk_level,
        "is_red_line_triggered": result.is_red_line_triggered
    }


@app.post("/api/v1/enterprises/batch-push", tags=["企业管理"], summary="批量按统一信用代码推送指标")
async def batch_push_indicators(request: Request):
    """
    批量推送接口，一次请求推送多家企业的指标数据。

    请求体格式：
    {
      "data_source": "数据中台",
      "records": [
        {"credit_code": "91440306XXXXXXXX", "indicators": {...}},
        ...
      ]
    }
    """
    body = await request.json()
    records = body.get('records', [])
    data_source = body.get('data_source', '批量推送')

    results = []
    for rec in records:
        credit_code = rec.get('credit_code', '').strip()
        indicator_data = rec.get('indicators', {})
        if not credit_code or not indicator_data:
            results.append({"credit_code": credit_code, "status": "skipped", "reason": "缺少必要字段"})
            continue
        ent = db.get_enterprise_by_credit_code(credit_code)
        if not ent:
            results.append({"credit_code": credit_code, "status": "not_found", "reason": "企业库中未找到该信用代码"})
            continue
        db.push_indicator_snapshot(credit_code, ent['enterprise_type'], indicator_data, data_source)
        result = calculate_risk(ent['enterprise_type'], indicator_data)
        details = [d.model_dump() for d in result.details]
        db.save_assessment(ent['id'], result.total_score, result.risk_level, result.is_red_line_triggered, details, result.recommended_actions, indicator_data)
        results.append({"credit_code": credit_code, "enterprise_name": ent['name'], "status": "success", "risk_score": result.total_score, "risk_level": result.risk_level})

    success_count = sum(1 for r in results if r['status'] == 'success')
    return {"message": f"批量推送完成，成功 {success_count}/{len(records)} 家", "results": results}


@app.get("/api/v1/enterprises/{eid}/history", tags=["企业管理"], summary="获取企业评估历史")
async def get_assessment_history(eid: int):
    """获取某企业的历史评估记录（最近20次）"""
    ent = db.get_enterprise(eid)
    if not ent:
        raise HTTPException(404, "企业不存在")
    history = db.get_assessment_history(ent['credit_code'])
    return {"enterprise": ent, "history": history}


@app.get("/api/v1/enterprises/{eid}/snapshots", tags=["企业管理"], summary="获取企业指标快照历史")
async def get_indicator_snapshots(eid: int):
    """获取某企业的历史指标快照（最近10次）"""
    ent = db.get_enterprise(eid)
    if not ent:
        raise HTTPException(404, "企业不存在")
    snapshots = db.list_snapshots(ent['credit_code'])
    return {"enterprise": ent, "snapshots": snapshots}

@app.get("/api/v1/enterprises/{eid}/profile", tags=["企业画像"], summary="获取企业全景风险画像数据")
async def get_enterprise_profile(eid: int):
    """获取企业360度风险画像，用于渲染图表和时间轴"""
    ent = db.get_enterprise(eid)
    if not ent:
        raise HTTPException(404, "企业不存在")
    
    # 获取最近 20 次评估历史用于绘制趋势图
    history = db.get_assessment_history(ent['credit_code'], limit=20)
    
    # 按照时间从小到大排序，方便前端绘图
    history = sorted(history, key=lambda x: x['assessed_at'])
    
    # 当前命中的维度分布（雷达图使用）
    radar_data = {}
    if ent.get('risk_details'):
        for detail in ent['risk_details']:
            cat = detail.get('category', '其他')
            radar_data[cat] = radar_data.get(cat, 0) + detail.get('score', 0)
            
    # 获取所有可能维度（保证雷达图结构完整）
    all_categories = set()
    config = load_config()
    common_rules = config.get('common', {}).get('rules', [])
    domain_rules = config.get(ent['enterprise_type'], {}).get('rules', [])
    for rule in common_rules + domain_rules:
        all_categories.add(rule.get('category', '其他'))
        
    radar = [{"name": c, "value": radar_data.get(c, 0)} for c in all_categories]

    # 构建规则名→规则定义的映射，用于补全指标明细中的字段值与阈值
    rule_map = {}
    for rule in common_rules + domain_rules:
        rule_map[rule.get('name', '')] = rule

    raw_data = ent.get('raw_data', {}) or {}
    indicator_details = []
    for detail in ent.get('risk_details', []):
        rule = rule_map.get(detail.get('item_name', ''), {})
        field = rule.get('field', '')
        actual_value = raw_data.get(field, None)
        indicator_details.append({
            **detail,
            'field': field,
            'operator': rule.get('operator', ''),
            'threshold': rule.get('threshold', ''),
            'actual_value': actual_value,
            'is_red_line': rule.get('is_red_line', False),
        })

    return {
        "enterprise": ent,
        "history": history,
        "radar": radar,
        "indicator_details": indicator_details
    }

@app.post("/api/v1/enterprises/batch-assess", tags=["企业管理"])
async def batch_assess():
    """批量评估所有有原始数据的企业"""
    all_ents = db.list_enterprises(page_size=9999)
    assessed = 0
    for ent in all_ents['items']:
        raw = ent.get('raw_data', {})
        if raw and isinstance(raw, dict) and len(raw) > 0:
            result = calculate_risk(ent['enterprise_type'], raw)
            details = [d.model_dump() for d in result.details]
            db.save_assessment(ent['id'], result.total_score, result.risk_level, result.is_red_line_triggered, details, result.recommended_actions, raw)
            assessed += 1
    return {"message": f"批量评估完成，共评估 {assessed} 家企业", "assessed": assessed}

# === 仪表盘 ===
@app.get("/api/v1/dashboard/summary", tags=["仪表盘"])
async def dashboard_summary():
    return db.get_dashboard_summary()

# === 指标总览 ===
@app.get("/api/v1/indicators", tags=["指标总览"])
async def list_indicators():
    """获取所有指标的触发统计概览"""
    return db.get_all_indicators()

@app.get("/api/v1/indicators/summary", tags=["指标总览"])
async def indicator_summary():
    """获取指标维度汇总统计"""
    return db.get_indicator_summary()

# === 指标导出 ===
@app.get("/api/v1/indicators/export", tags=["指标总览"])
async def export_indicators_xlsx():
    """导出所有预警指标的数据字段、规则与触发统计为Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    config = load_config()
    triggered = db.get_all_indicators()
    triggered_map = {ind['key']: ind.get('triggered_count', 0) for ind in triggered}

    wb = openpyxl.Workbook()

    # ── Sheet 1: 指标数据字段与规则 ──
    ws1 = wb.active
    ws1.title = "指标数据字段映射"

    hdr_font = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=10)
    hdr_fill = PatternFill(start_color='1A73E8', end_color='1A73E8', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='DADCE0'),
        right=Side(style='thin', color='DADCE0'),
        top=Side(style='thin', color='DADCE0'),
        bottom=Side(style='thin', color='DADCE0'),
    )

    def write_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

    s1_headers = ['领域', '指标标识', '指标名称', '维度分类', '五维分类', '扣分规则', '数据字段',
                  '运算符', '阈值', 'T1分值', 'T2阈值', 'T2分值', '是否红线', '指标说明', '触发企业数', '数据采集来源']
    write_header(ws1, s1_headers)

    domain_labels = {'common': '通用基础', 'factory': '制造业/服务业', 'construction': '工程建设'}
    red_fill = PatternFill(start_color='FFF0F0', end_color='FFF0F0', fill_type='solid')
    row = 2
    for domain in ['common', 'factory', 'construction']:
        domain_cfg = config.get(domain, {})
        for rule in domain_cfg.get('rules', []):
            is_red = rule.get('is_red_line', False)
            vals = [
                domain_labels.get(domain, domain),
                rule.get('key', ''),
                rule.get('name', ''),
                rule.get('category', ''),
                rule.get('five_category', ''),
                rule.get('scoring_rule', ''),
                rule.get('field', ''),
                rule.get('operator', ''),
                str(rule.get('threshold', '')),
                rule.get('score', 0),
                str(rule.get('threshold_t2', '') if rule.get('threshold_t2', '') is not None else ''),
                rule.get('score_t2', ''),
                '是' if is_red else '否',
                rule.get('description', ''),
                triggered_map.get(rule.get('key', ''), 0),
                rule.get('source', ''),
            ]
            for col, v in enumerate(vals, 1):
                cell = ws1.cell(row=row, column=col, value=v)
                cell.border = thin_border
                cell.font = Font(name='Microsoft YaHei', size=9)
                cell.alignment = Alignment(vertical='center', wrap_text=(col in (6, 14)))
                if is_red:
                    cell.fill = red_fill
            row += 1

    ws1.column_dimensions['A'].width = 14
    ws1.column_dimensions['B'].width = 28
    ws1.column_dimensions['C'].width = 30
    ws1.column_dimensions['D'].width = 22
    ws1.column_dimensions['E'].width = 14
    ws1.column_dimensions['F'].width = 40
    ws1.column_dimensions['G'].width = 28
    ws1.column_dimensions['H'].width = 8
    ws1.column_dimensions['I'].width = 12
    ws1.column_dimensions['J'].width = 8
    ws1.column_dimensions['K'].width = 12
    ws1.column_dimensions['L'].width = 8
    ws1.column_dimensions['M'].width = 10
    ws1.column_dimensions['N'].width = 50
    ws1.column_dimensions['O'].width = 12
    ws1.column_dimensions['P'].width = 28
    ws1.auto_filter.ref = f"A1:P{row - 1}"

    # ── Sheet 2: 维度汇总 ──
    ws2 = wb.create_sheet("维度分类汇总")
    s2_headers = ['领域', '维度名称', '指标数量', '红线指标数', '非红线最高分', '维度总分']
    write_header(ws2, s2_headers)

    category_map = {}
    for domain in ['common', 'factory', 'construction']:
        domain_cfg = config.get(domain, {})
        for rule in domain_cfg.get('rules', []):
            cat = rule.get('category', '未分类')
            k = (domain, cat)
            if k not in category_map:
                category_map[k] = {'count': 0, 'red_count': 0, 'max_score': 0, 'total_score': 0}
            category_map[k]['count'] += 1
            if rule.get('is_red_line', False):
                category_map[k]['red_count'] += 1
            else:
                s = max(rule.get('score', 0) or 0, rule.get('score_t2', 0) or 0)
                category_map[k]['max_score'] = max(category_map[k]['max_score'], s)
                category_map[k]['total_score'] += s

    row2 = 2
    for (domain, cat), stats in sorted(category_map.items()):
        vals = [domain_labels.get(domain, domain), cat, stats['count'],
                stats['red_count'], stats['max_score'], stats['total_score']]
        for col, v in enumerate(vals, 1):
            cell = ws2.cell(row=row2, column=col, value=v)
            cell.border = thin_border
            cell.font = Font(name='Microsoft YaHei', size=9)
        row2 += 1

    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 14
    ws2.column_dimensions['E'].width = 14
    ws2.column_dimensions['F'].width = 12
    ws2.auto_filter.ref = f"A1:F{row2 - 1}"

    # ── Sheet 3: 字段值对照表 ──
    ws3 = wb.create_sheet("字段值说明")
    s3_headers = ['字段分类', '字段名示例', '字段类型', '取值范围/说明']
    write_header(ws3, s3_headers)
    s3_data = [
        ['布尔类型', 'is_license_revoked / is_abnormal_business_status / has_recent_legal_rep_change …等', 'true / false', 'true=触发条件，false=正常'],
        ['百分比类型', 'capital_reduction_pct / equity_pledge_ratio / social_security_drop_pct …等', '整数（不带%号）', '如 30 表示30%，与阈值运算符比较'],
        ['数量类型', 'execution_case_count / penalty_count_1year / exception_count …等', '整数', '如 5 表示5次/5件'],
        ['天数/月数', 'months_until_term_end / months_since_establishment / days_in_abnormal_operation …等', '整数', '如 180 表示180天，6 表示6个月'],
        ['比率类型', 'paid_up_capital_ratio / controller_benefit_pct / branch_closure_ratio …等', '整数（不带%号）', '如 10 表示10%'],
    ]
    for i, row_data in enumerate(s3_data, 2):
        for col, v in enumerate(row_data, 1):
            cell = ws3.cell(row=i, column=col, value=v)
            cell.border = thin_border
            cell.font = Font(name='Microsoft YaHei', size=9)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    ws3.column_dimensions['A'].width = 16
    ws3.column_dimensions['B'].width = 50
    ws3.column_dimensions['C'].width = 22
    ws3.column_dimensions['D'].width = 55
    ws3.row_dimensions[1].height = 22
    for i in range(2, 7):
        ws3.row_dimensions[i].height = 40

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"wage_arrears_indicators_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*={filename}; filename={filename}"}
    )

# === 数据导出 ===
@app.get("/api/v1/enterprises/export/by-category", tags=["数据导出"])
async def export_by_indicator_category(category: str):
    """
    按预警指标维度分类导出企业数据（CSV格式）。
    category: 指标所属维度名称，如"基础经营与刚性支出健康度"
    """
    data = db.export_enterprises_by_indicator_category(category)

    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "分类风险数据"
    
    headers = ['企业名称', '统一社会信用代码', '企业类型', '所在街道', '所属行业',
               '员工数', '风险分数', '风险等级', '是否红线', '命中维度指标',
               '命中指标数', '最后评估时间', '建议处置']
    ws.append(headers)

    for row in data:
        matched_names = '；'.join([d.get('item_name', '') for d in row.get('matched_indicators', [])])
        ws.append([
            row.get('name', ''),
            row.get('credit_code', ''),
            row.get('enterprise_type', ''),
            row.get('district', ''),
            row.get('industry', ''),
            row.get('employee_count', ''),
            row.get('risk_score', ''),
            row.get('risk_level', ''),
            '是' if row.get('is_red_line') else '否',
            matched_names,
            row.get('matched_count', 0),
            row.get('last_assessed_at', ''),
            row.get('recommended_actions', '')
        ])

    safe_name = category.replace('/', '_').replace('\\', '_')
    # 移除所有非ASCII字符，确保文件名安全
    safe_name = ''.join(c if ord(c) < 128 else '_' for c in safe_name)
    safe_name = safe_name.strip('_') or 'indicator_category'
    filename = f"wage_arrears_category_{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# === 增强仪表盘 ===
@app.get("/api/v1/dashboard/enhanced", tags=["仪表盘"])
async def dashboard_enhanced():
    """增强仪表盘：月度趋势 + 街道分布 + 最新评估动态"""
    return db.get_enhanced_dashboard()


# === 高危指标排行 ===
@app.get("/api/v1/indicators/high-risk", tags=["指标总览"])
async def high_risk_indicators():
    """按触发企业数排序的高危指标排行"""
    return db.get_high_risk_indicators()


# === 指标定义列表 ===
@app.get("/api/v1/indicators/definitions", tags=["指标管理"])
async def indicator_definitions():
    """获取所有指标定义（从规则配置中提取）"""
    config = load_config()
    definitions = []
    domain_labels = {'common': '通用基础', 'factory': '制造业/服务业', 'construction': '工程建设'}
    for domain in ['common', 'factory', 'construction']:
        domain_cfg = config.get(domain, {})
        for rule in domain_cfg.get('rules', []):
            definitions.append({
                'key': rule.get('key', ''),
                'name': rule.get('name', ''),
                'domain': domain,
                'domain_name': domain_labels.get(domain, domain),
                'category': rule.get('category', ''),
                'field': rule.get('field', ''),
                'operator': rule.get('operator', ''),
                'threshold': rule.get('threshold', ''),
                'score': rule.get('score', 0),
                'five_category': rule.get('five_category', ''),
                'scoring_rule': rule.get('scoring_rule', ''),
                'threshold_t2': rule.get('threshold_t2', ''),
                'score_t2': rule.get('score_t2', None),
                'is_red_line': rule.get('is_red_line', False),
                'description': rule.get('description', ''),
                'source': rule.get('source', ''),
            })
    return definitions


# === 推送原始指标数据 (EAV格式) ===
@app.post("/api/v1/indicators/push-raw", tags=["指标管理"], summary="推送原始指标数据（EAV格式）")
async def push_raw_indicators(request: Request):
    """
    接收外部系统推送的原始指标数据（EAV格式），写入 raw_indicator_data，
    并尝试匹配企业库中的企业。

    请求体格式：
    {
      "data_source": "社保局",
      "records": [
        {"credit_code": "91440306XXXXXXXX", "indicator_key": "social_security_arrears", "value": true},
        ...
      ]
    }
    """
    body = await request.json()
    data_source = body.get('data_source', '外部推送')
    records = body.get('records', [])

    if not records:
        raise HTTPException(400, "records 不能为空")

    total = len(records)
    matched = 0
    unmatched = 0

    # Group by credit_code
    grouped = {}
    for rec in records:
        cc = rec.get('credit_code', '').strip()
        if not cc:
            continue
        if cc not in grouped:
            grouped[cc] = {}
        key = rec.get('indicator_key', '')
        value = rec.get('value')
        if key:
            grouped[cc][key] = value

    for credit_code, indicators in grouped.items():
        ent = db.get_enterprise_by_credit_code(credit_code)
        if ent:
            matched += 1
            # Write as indicator snapshot
            db.push_indicator_snapshot(credit_code, ent['enterprise_type'], indicators, data_source)
        else:
            unmatched += 1

    return {
        "message": f"原始指标推送完成",
        "status": "success",
        "total_records": total,
        "unique_enterprises": len(grouped),
        "matched": matched,
        "unmatched": unmatched,
        "data_source": data_source
    }


# === 数据库同步 ===
@app.get("/api/v1/admin/sync", tags=["管理维护"])
async def admin_sync():
    """数据库同步：确保表结构存在并种入默认数据"""
    try:
        db.init_db()
        db.seed_demo_data()
        return {"message": "数据库同步完成", "status": "success"}
    except Exception as e:
        raise HTTPException(500, f"同步失败: {str(e)}")


# === 模板下载 ===
@app.get("/api/v1/templates/enterprise-registry", tags=["模板下载"])
async def download_registry_template():
    """下载企业名册导入模板（Excel）"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "企业名册导入模板"

    headers = ['企业名称', '统一社会信用代码', '企业类型', '所在街道', '所属行业', '联系人', '联系电话', '在册员工数']
    hdr_font = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=10)
    hdr_fill = PatternFill(start_color='1A73E8', end_color='1A73E8', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='DADCE0'),
        right=Side(style='thin', color='DADCE0'),
        top=Side(style='thin', color='DADCE0'),
        bottom=Side(style='thin', color='DADCE0'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 示例数据
    example = ['示例企业名称', '91440306MA5FXXXXXX', 'factory', '新安街道', '制造业', '张三', '13800000000', 100]
    for col, v in enumerate(example, 1):
        cell = ws.cell(row=2, column=col, value=v)
        cell.font = Font(name='Microsoft YaHei', size=9)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 12

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=enterprise_registry_template.xlsx"}
    )


@app.get("/api/v1/templates/indicator-data", tags=["模板下载"])
async def download_indicator_template():
    """下载指标数据导入模板（Excel）"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "指标数据导入模板"

    hdr_font = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=10)
    hdr_fill = PatternFill(start_color='1A73E8', end_color='1A73E8', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='DADCE0'),
        right=Side(style='thin', color='DADCE0'),
        top=Side(style='thin', color='DADCE0'),
        bottom=Side(style='thin', color='DADCE0'),
    )

    # Sheet 1: 使用说明
    ws_note = wb.create_sheet("使用说明")
    notes = [
        ["字段名", "字段说明", "数据类型", "示例值"],
        ["credit_code", "统一社会信用代码（必填，用于匹配企业）", "文本", "91440306MA5F1234X1"],
        ["data_source", "数据来源标识", "文本", "社保局/税务局/法院/信访办"],
        ["is_license_revoked", "营业执照是否被吊销", "true/false", "false"],
        ["national_platform_weekly_complaints", "国家平台周投诉数", "整数", "0"],
        ["platform_complaints_weekly_count", "本地平台周投诉数", "整数", "2"],
        ["consecutive_water_arrears_cycles", "连续欠缴水费周期数", "整数", "1"],
        ["consecutive_electricity_arrears_cycles", "连续欠缴电费周期数", "整数", "3"],
        ["execution_case_count", "被执行人案件数", "整数", "2"],
        ["social_security_drop_pct", "社保参保人数下降百分比", "整数", "15"],
        ["industrial_power_drop_pct", "工业用电量下降百分比（仅制造业）", "整数", "10"],
        ["has_no_wage_account", "是否未开立工资专户（仅工程建设）", "true/false", "false"],
    ]
    for i, row_data in enumerate(notes, 1):
        for col, v in enumerate(row_data, 1):
            cell = ws_note.cell(row=i, column=col, value=v)
            if i == 1:
                cell.font = hdr_font
                cell.fill = hdr_fill
            else:
                cell.font = Font(name='Microsoft YaHei', size=9)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    ws_note.column_dimensions['A'].width = 35
    ws_note.column_dimensions['B'].width = 40
    ws_note.column_dimensions['C'].width = 18
    ws_note.column_dimensions['D'].width = 30

    # Sheet 2: 数据表
    headers = ['credit_code', 'data_source']
    # Add all indicator fields from rules_config.json
    indicator_fields = []
    config = load_config()
    for domain in ['common', 'factory', 'construction']:
        domain_cfg = config.get(domain, {})
        for rule in domain_cfg.get('rules', []):
            field = rule.get('field', '')
            if field and field not in indicator_fields:
                indicator_fields.append(field)
    headers.extend(indicator_fields)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # 示例数据行
    example_row = ['91440306MA5F1234X1', '社保局'] + [''] * len(indicator_fields)
    for col, v in enumerate(example_row, 1):
        cell = ws.cell(row=2, column=col, value=v)
        cell.font = Font(name='Microsoft YaHei', size=9)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    for col_idx in range(3, len(headers) + 1):
        ws.column_dimensions[chr(64 + min(col_idx, 90))].width = 14

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=indicator_data_template.xlsx"}
    )


# === 上传名册 ===
@app.post("/api/v1/enterprises/upload-registry", tags=["企业管理"])
async def upload_registry(request: Request):
    """上传企业名册 Excel 批量导入"""
    import openpyxl

    content_type = request.headers.get("content-type", "")
    if "multipart/form-data" not in content_type:
        raise HTTPException(400, "请使用 multipart/form-data 上传文件")

    form = await request.form()
    file = form.get("file")
    if not file:
        raise HTTPException(400, "未找到上传文件")

    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception as e:
        raise HTTPException(400, f"Excel 解析失败: {str(e)}")

    enterprises = []
    for row in rows:
        if not row[0] or not row[1]:
            continue
        enterprises.append({
            'name': str(row[0]).strip() if row[0] else '',
            'credit_code': str(row[1]).strip() if row[1] else '',
            'enterprise_type': str(row[2]).strip() if len(row) > 2 and row[2] else 'factory',
            'district': str(row[3]).strip() if len(row) > 3 and row[3] else '',
            'industry': str(row[4]).strip() if len(row) > 4 and row[4] else '',
            'contact_person': str(row[5]).strip() if len(row) > 5 and row[5] else '',
            'contact_phone': str(row[6]).strip() if len(row) > 6 and row[6] else '',
            'employee_count': int(row[7]) if len(row) > 7 and row[7] else 0,
        })

    if not enterprises:
        raise HTTPException(400, "未在文件中找到有效企业数据")

    created = db.batch_create_enterprises(enterprises)
    skipped = len(enterprises) - created
    return {
        "message": f"导入完成：成功 {created} 家" + (f"，跳过 {skipped} 家（可能信用代码重复）" if skipped > 0 else ""),
        "created": created,
        "skipped": skipped,
        "status": "success"
    }


# === 上传指标数据 ===
@app.post("/api/v1/enterprises/upload-indicators", tags=["企业管理"])
async def upload_indicators(request: Request):
    """上传企业指标数据 Excel 批量导入并自动评估"""
    import openpyxl

    content_type = request.headers.get("content-type", "")
    if "multipart/form-data" not in content_type:
        raise HTTPException(400, "请使用 multipart/form-data 上传文件")

    form = await request.form()
    file = form.get("file")
    if not file:
        raise HTTPException(400, "未找到上传文件")

    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        raise HTTPException(400, f"Excel 解析失败: {str(e)}")

    if len(rows) < 2:
        raise HTTPException(400, "文件中无数据行")

    headers = [str(h).strip() if h else '' for h in rows[0]]
    results = []
    for row in rows[1:]:
        credit_code = str(row[0]).strip() if row[0] else ''
        data_source = str(row[1]).strip() if len(row) > 1 and row[1] else 'Excel导入'
        if not credit_code:
            continue

        ent = db.get_enterprise_by_credit_code(credit_code)
        if not ent:
            results.append({"credit_code": credit_code, "status": "not_found", "reason": "企业库中未找到"})
            continue

        indicators = {}
        for i, h in enumerate(headers):
            if i < 2:
                continue
            if i < len(row) and row[i] is not None and str(row[i]).strip():
                val = str(row[i]).strip()
                if val.lower() == 'true':
                    indicators[h] = True
                elif val.lower() == 'false':
                    indicators[h] = False
                else:
                    try:
                        indicators[h] = int(val)
                    except ValueError:
                        try:
                            indicators[h] = float(val)
                        except ValueError:
                            indicators[h] = val

        db.push_indicator_snapshot(credit_code, ent['enterprise_type'], indicators, data_source)
        result = calculate_risk(ent['enterprise_type'], indicators)
        details = [d.model_dump() for d in result.details]
        db.save_assessment(ent['id'], result.total_score, result.risk_level, result.is_red_line_triggered, details, result.recommended_actions, indicators)
        results.append({
            "credit_code": credit_code,
            "enterprise_name": ent['name'],
            "status": "success",
            "risk_score": result.total_score,
            "risk_level": result.risk_level
        })

    success_count = sum(1 for r in results if r['status'] == 'success')
    return {
        "message": f"指标导入完成，成功评估 {success_count}/{len(results)} 家企业",
        "results": results,
        "status": "success"
    }


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 0))
    if not port:
        with open(CONFIG_PATH, 'r', encoding='utf-8') as f: cfg = json.load(f).get("system", {})
        port = cfg.get("api_port", 8000)
    uvicorn.run("main:app", host="0.0.0.0", port=port)
