"""
数据集导入脚本
将 F:\google agent\数据集\ 内的两个 Excel 文件导入到企业库并执行风险评估
运行方式: venv\Scripts\python import_dataset.py
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
import sqlite3
from datetime import datetime
from models import database as db
from models.scorer import calculate_risk

DATASET_DIR = r'F:\google agent\数据集'
ENTERPRISE_FILE = os.path.join(DATASET_DIR, '企业名单.xlsx')
INDICATOR_FILE  = os.path.join(DATASET_DIR, '企业信用风险指标.xlsx')

def load_enterprises():
    """读取企业名单，返回 {序号: {name, credit_code}} 字典"""
    wb = openpyxl.load_workbook(ENTERPRISE_FILE)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    ents = {}
    for r in range(2, ws.max_row + 1):
        row = {headers[c]: ws.cell(r, c + 1).value for c in range(len(headers))}
        # 兼容中文列名（编码后的列名也取第2/3列）
        seq  = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        code = ws.cell(r, 3).value
        if seq is None:
            continue
        ents[seq] = {'name': name, 'credit_code': str(code)}
    print(f"  读取到 {len(ents)} 家企业档案")
    return ents

def load_indicators():
    """读取风险指标数据，返回 {统一信用代码: {指标字段...}} 字典"""
    wb = openpyxl.load_workbook(INDICATOR_FILE)
    ws = wb.active
    # 第一行为列名
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    print(f"  指标列名({len(headers)}列): {headers}")
    
    # 找出指标字段（跳过序号/企业名/信用代码前3列）
    indicator_cols = headers[3:]  # 从第4列开始是指标
    
    data = {}
    for r in range(2, ws.max_row + 1):
        credit_code = str(ws.cell(r, 3).value)
        if not credit_code or credit_code == 'None':
            continue
        indicators = {}
        for i, col_name in enumerate(indicator_cols):
            val = ws.cell(r, 4 + i).value  # 从第4列开始
            if val is None:
                continue
            # 布尔型处理
            if isinstance(val, bool):
                indicators[col_name] = val
            elif isinstance(val, str) and val.upper() in ('TRUE', 'FALSE'):
                indicators[col_name] = val.upper() == 'TRUE'
            elif isinstance(val, (int, float)):
                indicators[col_name] = val
            else:
                indicators[col_name] = val
        data[credit_code] = indicators
    print(f"  读取到 {len(data)} 家企业的指标数据")
    return data

def main():
    print("=" * 60)
    print("  宝安区欠薪预警数据集导入工具")
    print("=" * 60)

    # 清空旧的模拟数据（仅清空，重新导入真实数据）
    conn = sqlite3.connect(db.DB_PATH)
    old_count = conn.execute("SELECT COUNT(*) FROM enterprises").fetchone()[0]
    print(f"\n[1/4] 当前数据库中已有 {old_count} 家企业")
    
    if old_count > 0:
        ans = input("  是否清空现有数据并重新导入？(y/n): ").strip().lower()
        if ans == 'y':
            conn.execute("DELETE FROM enterprises")
            conn.execute("DELETE FROM indicator_snapshots")
            conn.execute("DELETE FROM assessment_history")
            conn.commit()
            print("  已清空现有数据")
        else:
            print("  保留现有数据，将跳过已存在信用代码的企业")
    conn.close()

    print("\n[2/4] 读取企业名单...")
    enterprises = load_enterprises()

    print("\n[3/4] 读取风险指标数据...")
    indicators = load_indicators()

    print(f"\n[4/4] 导入企业并执行风险评估...")
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    success = 0
    skip = 0
    error = 0

    conn = sqlite3.connect(db.DB_PATH)
    
    for seq, ent in enterprises.items():
        name = ent['name']
        credit_code = ent['credit_code']
        
        # 检查是否已存在
        existing = conn.execute("SELECT id FROM enterprises WHERE credit_code=?", (credit_code,)).fetchone()
        if existing:
            skip += 1
            continue
        
        # 写入企业档案（全部默认为工厂类型，因为数据集只有工厂指标字段）
        try:
            conn.execute(
                "INSERT INTO enterprises (name,credit_code,enterprise_type,district,industry,created_at) VALUES (?,?,?,?,?,?)",
                (name, credit_code, 'factory', '', '', now)
            )
            conn.commit()
            eid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        except Exception as e:
            print(f"  ❌ 企业录入失败 [{name}]: {e}")
            error += 1
            continue

        # 匹配指标数据
        ind_data = indicators.get(credit_code, {})
        if not ind_data:
            print(f"  ⚠ 未找到 [{name}] 的指标数据，跳过评估")
            success += 1
            continue
        
        # 写入指标快照
        import json
        conn.execute(
            "INSERT INTO indicator_snapshots (credit_code,enterprise_type,indicator_data,data_source,snapshot_at) VALUES (?,?,?,?,?)",
            (credit_code, 'factory', json.dumps(ind_data, ensure_ascii=False), '数据集导入', now)
        )
        conn.commit()

        # 执行风险评估
        try:
            result = calculate_risk('factory', ind_data)
            details = [d.model_dump() for d in result.details]
            details_json = json.dumps(details, ensure_ascii=False)
            raw_json = json.dumps(ind_data, ensure_ascii=False)
            
            conn.execute(
                "UPDATE enterprises SET risk_score=?,risk_level=?,is_red_line=?,risk_details=?,recommended_actions=?,last_assessed_at=?,raw_data=? WHERE id=?",
                (result.total_score, result.risk_level, int(result.is_red_line_triggered), details_json, result.recommended_actions, now, raw_json, eid)
            )
            # 写入评估历史
            conn.execute(
                "INSERT INTO assessment_history (credit_code,enterprise_id,enterprise_name,enterprise_type,risk_score,risk_level,is_red_line,risk_details,recommended_actions,indicator_data,assessed_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                (credit_code, eid, name, 'factory', result.total_score, result.risk_level, int(result.is_red_line_triggered), details_json, result.recommended_actions, raw_json, now)
            )
            conn.commit()
            
            print(f"  [OK] [{name}] -> {result.risk_level} ({result.total_score}fen) {'[REDLINE!]' if result.is_red_line_triggered else ''}")
            success += 1
        except Exception as e:
            print(f"  [ERR] pinggu shibai [{name}]: {e}")
            error += 1

    conn.close()

    print("\n" + "=" * 60)
    print(f"  daoru wancheng! chenggong: {success} | tiaoguo: {skip} | shibai: {error}")
    print("=" * 60)

    # 打印统计摘要
    summary = db.get_dashboard_summary()
    print(f"\n📊 当前数据库风险分布:")
    print(f"  qiye zongsu: {summary['total']}  yipinggu: {summary['assessed']}")
    print(f"  [RED] hongse: {summary['red']}  [ORANGE] chengse: {summary['orange']}")
    print(f"  [YELLOW] huangse: {summary['yellow']}  [BLUE] lanse: {summary['blue']}")
    print(f"  [WARN] chufaxian: {summary['red_line_count']}")
    
    if summary['top_risk']:
        print(f"\n🏆 Top 5 高风险企业:")
        for i, e in enumerate(summary['top_risk'][:5]):
            print(f"  {i+1}. {e['name']} → {e['risk_level']} ({e['risk_score']}分)")

if __name__ == '__main__':
    main()
